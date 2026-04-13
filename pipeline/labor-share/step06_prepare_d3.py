"""
Step 6: Prepare D3-ready JSON datasets.

Produces three output files in data/processed/d3/:

1. d3_timeseries.json
   - Broad sectors (manufacturing, construction, services, others): 1990-2025
     reconstructed by anchoring 2008 levels and accumulating yearly changes.
   - Sub-sectors (11 services sub-sectors): 2009-2025
     derived from total employment (XLSX) minus resident employment (M182131).
   - GDP per industry where available.

2. d3_sectors.json
   - Sector metadata: key, label, level, parent, SSIC code, colour.

3. d3_overview.json
   - Economy-wide totals: total employed, resident, non-resident, overall share.
"""

import csv
import json
from pathlib import Path

import openpyxl
import pandas as pd

from src.config import (
    RAW_FILES, INTERIM_FILES, PROCESSED_FILES, MAPPINGS_DIR,
    ANALYSIS_START_YEAR, ANALYSIS_END_YEAR,
)

D3_DIR = PROCESSED_FILES["merged"].parent / "d3"

# ── Colours chosen to work for both light and dark backgrounds ──────────
SECTOR_COLOURS = {
    "manufacturing":      "#e45756",
    "construction":       "#f58518",
    "services":           "#4c78a8",
    "others":             "#9d9d9d",
    "wholesale_retail":   "#54a24b",
    "transport_storage":  "#72b7b2",
    "accomm_food":        "#b279a2",
    "info_comms":         "#eeca3b",
    "finance_insurance":  "#ff9da6",
    "real_estate":        "#9ecae9",
    "professional":       "#d67195",
    "admin_support":      "#b07aa1",
    "community_social":   "#79706e",
    "total":              "#333333",
}


# ─────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────

def _load_changes() -> dict:
    """Return {(year, industry, res_stat): change_thousands}."""
    ch = {}
    with open(RAW_FILES["emp_change_by_res"]) as f:
        for r in csv.DictReader(f):
            try:
                val = float(r["employment_change"])
            except (ValueError, TypeError):
                val = 0.0
            ch[(int(r["year"]), r["industry1"], r["res_stat"])] = val
    return ch


def _load_2008_anchors() -> dict:
    """Return {sector: {resident, nonresident, total}} for 2008."""
    # Total employment from XLSX SSIC2020 sheet
    wb = openpyxl.load_workbook(RAW_FILES["total_emp_level"], data_only=True)
    ws = wb["SSIC2020"]
    headers = []
    for cell in ws[3]:
        v = cell.value
        if v and hasattr(v, "year"):
            headers.append(v.year)
        elif v and str(v)[:4].isdigit():
            headers.append(int(str(v)[:4]))
        else:
            headers.append(None)

    name_map = {
        "Total": "total", "Manufacturing": "manufacturing",
        "Construction": "construction", "Services": "services",
    }
    totals = {}
    for row in ws.iter_rows(min_row=4, max_row=56):
        name = str(row[1].value).strip() if row[1].value else ""
        ssic = str(row[0].value).strip() if row[0].value else ""
        key = name_map.get(name) or (("others") if ssic == "A, B, D, E" else None)
        if key:
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i] == 2008 and cell.value is not None:
                    totals[key] = float(cell.value)

    # Resident employment from M182131
    with open(RAW_FILES["resident_emp"]) as f:
        data = json.load(f)
    res_map = {
        "All Industries (Employed Residents)": "total",
        "Manufacturing": "manufacturing",
        "Construction": "construction",
        "Services": "services",
        "Others": "others",
    }
    residents = {}
    for r in data["Data"]["row"][:17]:
        name = r["rowText"].strip()
        sk = res_map.get(name)
        if sk:
            for c in r["columns"]:
                if c["key"] == "2008" and c.get("value") and c["value"] != "na":
                    residents[sk] = float(c["value"])

    anchors = {}
    for sk in ["total", "manufacturing", "construction", "services", "others"]:
        t = totals.get(sk, 0)
        res = residents.get(sk, 0)
        anchors[sk] = {"resident": res, "nonresident": t - res, "total": t}
    return anchors


# ─────────────────────────────────────────────────────────────────────────
# 1. Build broad-sector time series 1990-2025
# ─────────────────────────────────────────────────────────────────────────

def build_broad_timeseries() -> list[dict]:
    """Build level-1 sector time series.

    2009-2025: use DIRECT derivation (XLSX total − M182131 resident)
               from the step 4/5 pipeline output — most accurate.
    1990-2008: RECONSTRUCT by anchoring 2009 direct values and
               walking backward using MOM yearly employment changes.
               Note: pre-2008 totals may not match XLSX exactly due
               to SSIC classification breaks (1996→2005→2020).
    """
    ch = _load_changes()
    sectors = ["manufacturing", "construction", "services", "others"]

    # ── Load 2009-2025 direct derivation from step 5 output ──
    direct_df = pd.read_csv(PROCESSED_FILES["foreign_share"])
    direct = {}  # {(year, sector): {resident, nonresident, total}}
    for _, row in direct_df[direct_df["level"] == 1].iterrows():
        sk = row["sector_key"]
        yr = int(row["year"])
        if sk in sectors and pd.notna(row.get("total_employment_thousands")):
            res = row.get("resident_employment_thousands")
            nr = row.get("nonresident_employment_thousands")
            if pd.notna(res) and pd.notna(nr):
                direct[(yr, sk)] = {
                    "resident": round(float(res), 2),
                    "nonresident": round(float(nr), 2),
                    "total": round(float(row["total_employment_thousands"]), 2),
                }

    all_records = []

    for sk in sectors:
        # ── 2009-2025: direct values ──
        for year in range(2009, 2026):
            d = direct.get((year, sk))
            if d:
                total = d["total"]
                share = round(d["nonresident"] / total * 100, 2) if total > 0 else None
                all_records.append({
                    "year": year,
                    "sector_key": sk,
                    "level": 1,
                    "total_emp": total,
                    "resident_emp": d["resident"],
                    "nonresident_emp": d["nonresident"],
                    "nonresident_share_pct": share,
                    "data_source": "direct",
                })

        # ── 1990-2008: reconstruct backward from 2009 anchor ──
        anchor = direct.get((2009, sk))
        if not anchor:
            continue
        res_l, nr_l = anchor["resident"], anchor["nonresident"]
        for year in range(2009, 1990, -1):
            # Undo change for `year` to get end-of-(year-1) level
            res_ch = ch.get((year, sk, "resident"), 0)
            nr_ch = ch.get((year, sk, "non-resident"), 0)
            res_l -= res_ch
            nr_l -= nr_ch
            total = round(res_l + nr_l, 2)
            share = round(nr_l / total * 100, 2) if total > 0 else None
            all_records.append({
                "year": year - 1,
                "sector_key": sk,
                "level": 1,
                "total_emp": total,
                "resident_emp": round(res_l, 2),
                "nonresident_emp": round(nr_l, 2),
                "nonresident_share_pct": share,
                "data_source": "reconstructed",
            })

    return all_records


# ─────────────────────────────────────────────────────────────────────────
# 2. Build sub-sector time series 2009-2025
# ─────────────────────────────────────────────────────────────────────────

def build_subsector_timeseries() -> list[dict]:
    """Use processed merged data for level-2 sub-sectors."""
    df = pd.read_csv(PROCESSED_FILES["foreign_share"])
    sub = df[df["level"] == 2].copy()

    # Exclude sectors with persistent data quality issues
    # finance_insurance has negative derived non-resident in most years
    flagged = sub.groupby("sector_key").apply(
        lambda g: (g["nonresident_employment_thousands"] < 0).sum() / len(g)
    )
    bad_sectors = flagged[flagged > 0.4].index.tolist()
    if bad_sectors:
        print(f"  [INFO] Excluding sub-sectors with >40% flagged rows: {bad_sectors}")
        sub = sub[~sub["sector_key"].isin(bad_sectors)]

    records = []
    for _, row in sub.iterrows():
        if pd.isna(row.get("total_employment_thousands")):
            continue
        total = row["total_employment_thousands"]
        resident = row.get("resident_employment_thousands")
        nonresident = row.get("nonresident_employment_thousands")
        share = row.get("nonresident_share_pct")
        records.append({
            "year": int(row["year"]),
            "sector_key": row["sector_key"],
            "level": 2,
            "total_emp": round(total, 2) if pd.notna(total) else None,
            "resident_emp": round(resident, 2) if pd.notna(resident) else None,
            "nonresident_emp": round(nonresident, 2) if pd.notna(nonresident) else None,
            "nonresident_share_pct": round(share, 2) if pd.notna(share) else None,
            "data_source": "direct",
        })
    return records


# ─────────────────────────────────────────────────────────────────────────
# 3. Add GDP data
# ─────────────────────────────────────────────────────────────────────────

def attach_gdp(records: list[dict]) -> list[dict]:
    """Attach GDP values to records using the interim GDP CSV."""
    gdp_df = pd.read_csv(INTERIM_FILES["gdp"])
    gdp_lookup = {}
    for _, row in gdp_df.iterrows():
        gdp_lookup[(int(row["year"]), row["sector_key"])] = row["gdp_current_million_sgd"]

    for rec in records:
        key = (rec["year"], rec["sector_key"])
        gdp_val = gdp_lookup.get(key)
        rec["gdp_million_sgd"] = round(gdp_val, 2) if gdp_val and not pd.isna(gdp_val) else None
        if rec.get("total_emp") and gdp_val and not pd.isna(gdp_val) and rec["total_emp"] > 0:
            # million SGD ÷ thousands of workers = thousand SGD per worker
            rec["gdp_per_worker_thousand_sgd"] = round(
                gdp_val / rec["total_emp"], 2
            )
        else:
            rec["gdp_per_worker_thousand_sgd"] = None

    return records


# ─────────────────────────────────────────────────────────────────────────
# 4. Build economy-wide overview
# ─────────────────────────────────────────────────────────────────────────

def build_overview(records: list[dict]) -> list[dict]:
    """Aggregate level-1 sectors into economy-wide totals per year."""
    from collections import defaultdict
    yearly = defaultdict(lambda: {"resident": 0, "nonresident": 0})
    for rec in records:
        if rec["level"] == 1:
            yearly[rec["year"]]["resident"] += rec.get("resident_emp") or 0
            yearly[rec["year"]]["nonresident"] += rec.get("nonresident_emp") or 0

    overview = []
    for year in sorted(yearly):
        d = yearly[year]
        total = round(d["resident"] + d["nonresident"], 2)
        share = round(d["nonresident"] / total * 100, 2) if total > 0 else None
        overview.append({
            "year": year,
            "total_emp": total,
            "resident_emp": round(d["resident"], 2),
            "nonresident_emp": round(d["nonresident"], 2),
            "nonresident_share_pct": share,
        })
    return overview


# ─────────────────────────────────────────────────────────────────────────
# 5. Build sector metadata
# ─────────────────────────────────────────────────────────────────────────

def build_sector_meta(records: list[dict]) -> list[dict]:
    mapping = pd.read_csv(MAPPINGS_DIR / "industry_code_mapping.csv")
    seen_keys = set(r["sector_key"] for r in records)

    LABELS = {
        "manufacturing": "Manufacturing",
        "construction": "Construction",
        "services": "Services",
        "others": "Others (Agri, Mining, Utilities)",
        "wholesale_retail": "Wholesale & Retail Trade",
        "transport_storage": "Transportation & Storage",
        "accomm_food": "Accommodation & Food Services",
        "info_comms": "Information & Communications",
        "finance_insurance": "Financial & Insurance Services",
        "real_estate": "Real Estate Services",
        "professional": "Professional Services",
        "admin_support": "Administrative & Support Services",
        "community_social": "Community, Social & Personal Services",
    }

    meta = []
    for _, row in mapping.iterrows():
        sk = row["sector_key"]
        if sk not in seen_keys:
            continue
        meta.append({
            "sector_key": sk,
            "label": LABELS.get(sk, sk.replace("_", " ").title()),
            "level": int(row["level"]),
            "parent": row["parent_sector_key"] if pd.notna(row["parent_sector_key"]) else None,
            "ssic_2020_code": row["ssic_2020_code"] if pd.notna(row["ssic_2020_code"]) else None,
            "colour": SECTOR_COLOURS.get(sk, "#888888"),
        })
    return meta


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────

def main():
    print("Step 6: Preparing D3-ready datasets...")
    D3_DIR.mkdir(parents=True, exist_ok=True)

    # Broad sectors 1990-2025
    broad = build_broad_timeseries()
    print(f"  Broad sectors: {len(broad)} rows (1990-2025, 4 sectors)")

    # Sub-sectors 2009-2025
    sub = build_subsector_timeseries()
    print(f"  Sub-sectors:   {len(sub)} rows (2009-2025)")

    all_records = broad + sub
    all_records = attach_gdp(all_records)
    all_records.sort(key=lambda r: (r["year"], r["level"], r["sector_key"]))

    # Write timeseries (JSON + CSV)
    with open(D3_DIR / "d3_timeseries.json", "w") as f:
        json.dump(all_records, f, indent=2)
    ts_cols = [
        "year", "sector_key", "level", "total_emp", "resident_emp",
        "nonresident_emp", "nonresident_share_pct", "gdp_million_sgd",
        "gdp_per_worker_thousand_sgd", "data_source",
    ]
    _write_csv(D3_DIR / "d3_timeseries.csv", all_records, ts_cols)
    print(f"  → d3_timeseries.json/csv ({len(all_records)} records)")

    # Sector metadata (JSON + CSV)
    meta = build_sector_meta(all_records)
    with open(D3_DIR / "d3_sectors.json", "w") as f:
        json.dump(meta, f, indent=2)
    _write_csv(D3_DIR / "d3_sectors.csv", meta,
               ["sector_key", "label", "level", "parent", "ssic_2020_code", "colour"])
    print(f"  → d3_sectors.json/csv ({len(meta)} sectors)")

    # Economy overview (JSON + CSV)
    overview = build_overview(all_records)
    with open(D3_DIR / "d3_overview.json", "w") as f:
        json.dump(overview, f, indent=2)
    _write_csv(D3_DIR / "d3_overview.csv", overview,
               ["year", "total_emp", "resident_emp", "nonresident_emp", "nonresident_share_pct"])
    print(f"  → d3_overview.json/csv ({len(overview)} years)")

    return all_records, meta, overview


def _write_csv(path: Path, records: list[dict], fieldnames: list[str]):
    import csv as csv_mod
    with open(path, "w", newline="") as f:
        w = csv_mod.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(records)


if __name__ == "__main__":
    import sys, os
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
    main()
