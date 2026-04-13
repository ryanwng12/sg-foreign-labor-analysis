"""
Step 1: Standardize total employment levels by industry from MOM XLSX.

Reads the SSIC2020 sheet from mom_emp_level_by_ind.xlsx and outputs a long-format
CSV with columns: year, sector_key, total_employment_thousands.
"""

import pandas as pd
import openpyxl
from src.config import RAW_FILES, INTERIM_FILES, MAPPINGS_DIR

XLSX_NAME_TO_SECTOR_KEY = {}


def _load_mapping():
    """Build lookup from XLSX industry name → sector_key."""
    global XLSX_NAME_TO_SECTOR_KEY
    mapping = pd.read_csv(MAPPINGS_DIR / "industry_code_mapping.csv")
    for _, row in mapping.iterrows():
        name = row["total_emp_xlsx_name"]
        if pd.notna(name) and name.strip():
            XLSX_NAME_TO_SECTOR_KEY[name.strip()] = row["sector_key"]


def extract_ssic2020_sheet() -> pd.DataFrame:
    """Parse SSIC2020 sheet into long-format DataFrame."""
    wb = openpyxl.load_workbook(RAW_FILES["total_emp_level"], data_only=True)
    ws = wb["SSIC2020"]

    # Row 3 contains year headers (datetime or string)
    headers = []
    for cell in ws[3]:
        v = cell.value
        if v is None:
            headers.append(None)
        elif hasattr(v, "year"):
            headers.append(v.year)
        else:
            s = str(v)
            if s[:4].isdigit():
                headers.append(int(s[:4]))
            else:
                headers.append(None)

    records = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        industry_name = str(row[1].value).strip() if row[1].value else None
        if not industry_name:
            continue
        sector_key = XLSX_NAME_TO_SECTOR_KEY.get(industry_name)
        if sector_key is None:
            print(f"  [WARN] No mapping for XLSX industry: '{industry_name}'")
            continue
        for i, cell in enumerate(row):
            year = headers[i] if i < len(headers) else None
            if year and isinstance(year, int) and cell.value is not None:
                try:
                    records.append({
                        "year": year,
                        "sector_key": sector_key,
                        "total_employment_thousands": float(cell.value),
                    })
                except (ValueError, TypeError):
                    pass

    return pd.DataFrame(records)


def main():
    print("Step 1: Standardizing total employment levels (XLSX SSIC2020)...")
    _load_mapping()
    df = extract_ssic2020_sheet()
    df = df.sort_values(["sector_key", "year"]).reset_index(drop=True)
    df.to_csv(INTERIM_FILES["total_emp"], index=False)
    print(f"  Wrote {len(df)} rows ({df['sector_key'].nunique()} sectors, "
          f"{df['year'].min()}-{df['year'].max()}) → {INTERIM_FILES['total_emp'].name}")
    return df


if __name__ == "__main__":
    import sys, os
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
    main()
